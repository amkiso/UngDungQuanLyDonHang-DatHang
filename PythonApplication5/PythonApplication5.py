
# -*- coding: utf-8 -*-
"""
Created on Tue Apr 01 10:04:00 2025

@author: HOÀNG VŨ
TÁC GIẢ: 2001231059-Lô Hoàng Vũ, stt:50
SỐ THÀNH VIÊN THỰC HIỆN: 01
NGÀY HOÀN THÀNH: 27/05/2025
*Các hàm, biến thường xuyên được đặt bằng tiếng Anh vì thông dụng, dễ nhớ và ngắn gọn hơn Tiếng Việt!*
"""

from ast import Lambda
from enum import global_enum
import tkinter as tk
import base64
import re
from tkinter import Scrollbar, ttk, messagebox
import json
import datetime
import hashlib
import os,sys
import traceback
from tkinter import filedialog
from turtle import reset
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
__version__ = "1.0.1"

def resource_path(relative_path):
   
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

class DonHang:
    def __init__(self, Ma, TenH, NguoiGui, NNhan, cannang, DVVC, status, From, TO, Pnumber1, Pnumber2, COD, NgayDatHang=None):
        self.Ma = Ma
        self.TenH = TenH
        self.From = From
        self.TO = TO
        self.cannang = cannang
        self.DVVC = DVVC
        self.status = status
        self.Pnumber1 = Pnumber1
        self.Pnumber2 = Pnumber2
        self.NguoiGui = NguoiGui
        self.NNhan = NNhan
        self.COD = COD
        self.NgayDatHang = NgayDatHang if NgayDatHang else datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    
    def to_dict(self):
        return {
            "Mã đơn hàng": self.Ma,
            "Tên đơn hàng": self.TenH,
            "Tên shop": self.NguoiGui,
            "Tên người nhận": self.NNhan,
            "Cân nặng (kg)": self.cannang,
            "Đơn vị vận chuyển": self.DVVC,
            "Trạng thái": self.status,
            "Địa chỉ nhận hàng": self.TO,
            "Địa chỉ lấy hàng": self.From,
            "SĐT người nhận": self.Pnumber1,
            "SĐT người gửi": self.Pnumber2,
            "COD": self.COD,
            "Ngày đặt hàng": self.NgayDatHang
        }
class QuanLyDonHang:
    def __init__(self, FileName):
        self.FileName = FileName
    
    def FileRead(self):
        try:
            with open(self.FileName, "r", encoding="utf-8") as file:
               try:
                data = json.load(file)
               except json.JSONDecodeError:
                   return []
               if not isinstance(data, list):
                    return []
               if not data:
                   return []
               orders = []
               for d in data:
                    orders.append(DonHang(
                        d["Mã đơn hàng"], d["Tên đơn hàng"], d["Tên shop"], d["Tên người nhận"],
                        d["Cân nặng (kg)"], d["Đơn vị vận chuyển"], d["Trạng thái"], d["Địa chỉ lấy hàng"],
                        d["Địa chỉ nhận hàng"], d["SĐT người nhận"], d["SĐT người gửi"],
                        d["COD"], d["Ngày đặt hàng"]
                    ))
               return orders
        except (FileNotFoundError, json.JSONDecodeError, KeyError, ValueError) as e:
            messagebox.showwarning("Lỗi", f"Không thể đọc dữ liệu: {e}")
            return []
    
    def save_orders(self, orders):
        with open(self.FileName, "w", encoding="utf-8") as file:
            json.dump([order.to_dict() for order in orders], file, indent=4, ensure_ascii=False)
#2 hàm kiểm tra focus
def On_entry_click(event, entry, default_text):
    if entry.get() == default_text:
        entry.delete(0, "end")
        entry.config(fg="black")
def On_focus_out(event, entry, default_text):
    if entry.get().strip() == "":
        entry.insert(0, default_text)
        entry.config(fg="gray")
#chi tiết đơn hàng
def show_order_details(order):
    detail_window = tk.Toplevel()
    detail_window.iconbitmap(resource_path("ico\OIP.ico")) 
    detail_window.title("Chi tiết đơn hàng")
    detail_window.geometry("700x350")
    
    details = order.to_dict()
    text = ""
    current_width = 0
    
    for key, value in details.items():
        item_text = f"{key}: {value}    "
        if current_width + len(item_text) * 7 > 600:
            text += "\n\n"
            current_width = 0
        text += item_text
        current_width += len(item_text) * 7
    
    tk.Label(detail_window, text=text, justify="left", wraplength=600, font=("Arial", 13), padx=10, pady=10).pack()
def show_selected_order():
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Lỗi", "Vui lòng chọn một đơn hàng!")
        return
    
    values = tree.item(selected_item, "values")
    orders = order_manager.FileRead()
    
    for order in orders:
        if order.Ma == values[1]:
            show_order_details(order)
            break
#kiểm tra định dạng email
def is_valid_email(email):
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'
    return re.match(pattern, email) is not None
#Lưu user
def save_user(user):
    users = load_users()
    # Mã hóa mật khẩu trước khi lưu
    user['password'] = hash_password(user['password'])
    users.append(user)
    with open("User.json", "w", encoding="utf-8") as f:
        json.dump(users, f, indent=4, ensure_ascii=False)
#xóa đơn hàng
def delete_selected_order():
    selected_item = tree.selection()
    
    if not selected_item:
        messagebox.showwarning("Lỗi", "Vui lòng chọn một đơn hàng để xóa!")
        return

    values = tree.item(selected_item, "values")
    order_id = values[1]

    confirm = messagebox.askyesno("Xác nhận xóa", f"Bạn có chắc muốn xóa đơn hàng '{order_id}' không?")
    if not confirm:
        return

    orders = order_manager.FileRead()
    new_orders = [order for order in orders if order.Ma != order_id]
    order_manager.save_orders(new_orders)
    load_orders()
    messagebox.showinfo("Thành công", "Đơn hàng đã được xóa thành công!")
#thống kê
def ThongKe():
    if current_file is None:
        messagebox.showwarning("Thông báo", "Chưa có dữ liệu để thống kê.")
        return
    QT, VN = 0, 0
    tong_khoi_luong = 0.0
    khoi_luong_kho = 0.0
    khoi_luong_van_chuyen = 0.0
    khoi_luong_da_giao = 0.0
    doanh_thu = 0
    loi_nhuan = 0
    tong_don = 0
    da_giao = 0
    dang_giao = 0
    dang_van_chuyen = 0
    da_huy = 0
    trong_kho = 0

    sp_stats = {"DT": 0, "GD": 0, "TT": 0, "TR": 0, "LT": 0}

    orders = order_manager.FileRead()
    tong_don = len(orders)

    for order in orders:
        ma_don = order.Ma
        loai_don = ma_don[:2]
        loai_sp = ma_don[2:4]

        if loai_don == "QT":
            QT += 1
        elif loai_don == "VN":
            VN += 1

        if loai_sp in sp_stats:
            sp_stats[loai_sp] += 1

        kl = float(order.cannang)
        tong_khoi_luong += kl

        if order.status == "Đã nhập kho":
            khoi_luong_kho += kl
            trong_kho += 1
        elif order.status == "Đang vận chuyển":
            khoi_luong_van_chuyen += kl
            dang_van_chuyen += 1
        elif order.status == "Đã giao":
            khoi_luong_da_giao += kl
            da_giao += 1
        elif order.status == "Đang giao":
            dang_giao += 1
        elif order.status == "Đã hủy":
            da_huy += 1

        if kl < 1:
            doanh_thu += 16500
        elif 1 <= kl < 2:
            doanh_thu += 21800
        elif 2 <= kl < 3:
            doanh_thu += 31000
        else:
            doanh_thu += 31000 + int((kl - 3) / 0.5) * 5750

        if hasattr(order, "COD") and order.COD == "1":
            doanh_thu += 10000

    loi_nhuan = int(doanh_thu * 0.3)
    ti_le_thanh_cong = (da_giao / tong_don * 100) if tong_don > 0 else 0
    ti_le_VN = (VN / tong_don * 100) if tong_don > 0 else 0
    ti_le_QT = (QT / tong_don * 100) if tong_don > 0 else 0
    # Tránh chia 0
    ti_le_DT = sp_stats["DT"] / tong_don * 100 if tong_don else 0
    ti_le_GD = sp_stats["GD"] / tong_don * 100 if tong_don else 0
    ti_le_TT = sp_stats["TT"] / tong_don * 100 if tong_don else 0
    ti_le_TR = sp_stats["TR"] / tong_don * 100 if tong_don else 0
    ti_le_LT = sp_stats["LT"] / tong_don * 100 if tong_don else 0

    ti_le_kg_giao = khoi_luong_da_giao / tong_khoi_luong * 100 if tong_khoi_luong else 0
    ti_le_kg_vc = khoi_luong_van_chuyen / tong_khoi_luong * 100 if tong_khoi_luong else 0
    ti_le_kg_kho = khoi_luong_kho / tong_khoi_luong * 100 if tong_khoi_luong else 0

    ti_le_loi_nhuan = loi_nhuan / doanh_thu * 100 if doanh_thu else 0

# Danh sách thống kê đầy đủ
    thong_ke = [
    ("Tổng số đơn hàng", tong_don, "100%"),
    ("Đơn nội địa (VN)", VN, f"{ti_le_VN:.1f}%"),
    ("Đơn quốc tế (QT)", QT, f"{ti_le_QT:.1f}%"),

    ("Đã giao", da_giao, f"{da_giao / tong_don * 100:.1f}%" if tong_don else ""),
    ("Đang giao", dang_giao, f"{dang_giao / tong_don * 100:.1f}%" if tong_don else ""),
    ("Đang vận chuyển", dang_van_chuyen, f"{dang_van_chuyen / tong_don * 100:.1f}%" if tong_don else ""),
    ("Trong kho", trong_kho, f"{trong_kho / tong_don * 100:.1f}%" if tong_don else ""),
    ("Đã hủy", da_huy, f"{da_huy / tong_don * 100:.1f}%" if tong_don else ""),

    ("Tỉ lệ thành công", "-",f"{ti_le_thanh_cong:.1f}%"),

    ("Điện tử (DT)", sp_stats["DT"], f"{ti_le_DT:.1f}%"),
    ("Gia dụng (GD)", sp_stats["GD"], f"{ti_le_GD:.1f}%"),
    ("Thể thao (TT)", sp_stats["TT"], f"{ti_le_TT:.1f}%"),
    ("Thời trang (TR)", sp_stats["TR"], f"{ti_le_TR:.1f}%"),
    ("Linh tinh (LT)", sp_stats["LT"], f"{ti_le_LT:.1f}%"),

    ("Tổng khối lượng (kg)", f"{tong_khoi_luong:.2f}", "100%"),
    ("Đã giao (kg)", f"{khoi_luong_da_giao:.2f}", f"{ti_le_kg_giao:.1f}%"),
    ("Đang vận chuyển (kg)", f"{khoi_luong_van_chuyen:.2f}", f"{ti_le_kg_vc:.1f}%"),
    ("Trong kho (kg)", f"{khoi_luong_kho:.2f}", f"{ti_le_kg_kho:.1f}%"),

    ("Doanh thu (VNĐ)", f"{doanh_thu:,}", "100%"),
    ("Lợi nhuận (30%)", f"{loi_nhuan:,}", f"{ti_le_loi_nhuan:.1f}%"),
]



    window = tk.Toplevel()
    window.title("Thống Kê Đơn Hàng")
    window.geometry("750x500")
    window.configure(bg="#ECEFF1")
    window.iconbitmap(resource_path("ico\chart.ico"))
    label = tk.Label(window, text="BẢNG THỐNG KÊ ĐƠN HÀNG", 
                     font=("Arial", 16, "bold"), fg="#1A237E", bg="#ECEFF1")
    label.pack(pady=10)

    # Treeview
    tree = ttk.Treeview(window, columns=("ChiTieu", "GiaTri", "ChiTiet"), 
                        show="headings", height=20)
    tree.heading("ChiTieu", text="Thông tin")
    tree.heading("GiaTri", text="Giá trị")
    tree.heading("ChiTiet", text="Tỉ lệ")

    tree.column("ChiTieu", width=250)
    tree.column("GiaTri", width=120, anchor="center")
    tree.column("ChiTiet", width=150, anchor="center")

    tree.tag_configure("evenrow", background="#f0f0f0")  # xám sáng
    tree.tag_configure("oddrow", background="#ffffff")   # trắng

    # Insert dữ liệu với tag
    for idx, item in enumerate(thong_ke):
        tag = 'evenrow' if idx % 2 == 0 else 'oddrow'
        tree.insert("", "end", values=item, tags=(tag,))

    # Scrollbar
    vsb = ttk.Scrollbar(window, orient="vertical", command=tree.yview)
    tree.configure(yscrollcommand=vsb.set)
    vsb.pack(side="right", fill="y")

    tree.pack(expand=True, fill="both", padx=20, pady=10)

    tk.Button(window, text="Đóng", command=window.destroy,
              font=("Arial", 12), bg="#EF5350", fg="white").pack(pady=10)
#bộ lọc
def Loc_San_Pham(parent_window, tree, order_manager):
   if current_file is None:
        messagebox.showwarning("Thông báo", "Bạn chưa có file để áp dụng bộ lọc.")
        return

   Loc = tk.Toplevel(parent_window)
   Loc.title("Lọc Đơn Hàng")
   Loc.geometry("500x250")
   Loc.configure(bg="lightblue")
   Loc.iconbitmap(resource_path("ico/FindS.ico"))

    # Labels
   tk.Label(Loc, text="Loại đơn", bg="lightblue").grid(row=0, column=0, sticky='w', padx=10, pady=5)
   tk.Label(Loc, text="Loại sản phẩm", bg="lightblue").grid(row=1, column=0, sticky='w', padx=10, pady=5)
   tk.Label(Loc, text="Phương thức giao hàng", bg="lightblue").grid(row=2, column=0, sticky='w', padx=10, pady=5)
   tk.Label(Loc, text="Trạng thái", bg="lightblue").grid(row=3, column=0, sticky='w', padx=10, pady=5)
   tk.Label(Loc, text="Cân nặng", bg="lightblue").grid(row=4, column=0, sticky='w', padx=10, pady=5)
   tk.Label(Loc, text="Thời gian", bg="lightblue").grid(row=5, column=0, sticky='w', padx=10, pady=5)

    # Comboboxes
   loai_don_cb = ttk.Combobox(Loc, values=["", "Quốc tế", "Nội địa"], state="readonly")
   loai_san_pham_cb = ttk.Combobox(Loc, values=["", "Điện tử", "Gia dụng", "Thể thao", "Thời trang", "Khác"], state="readonly")
   phuong_thuc_cb = ttk.Combobox(Loc, values=["", "COD", "Non-COD"], state="readonly")
   trang_thai_cb = ttk.Combobox(Loc, values=["", "Đã giao", "Đang giao", "Đang vận chuyển", "Đã nhập kho", "Đã hủy"], state="readonly")

   loai_don_cb.grid(row=0, column=1, columnspan=2, sticky='we', padx=5)
   loai_san_pham_cb.grid(row=1, column=1, columnspan=2, sticky='we', padx=5)
   phuong_thuc_cb.grid(row=2, column=1, columnspan=2, sticky='we', padx=5)
   trang_thai_cb.grid(row=3, column=1, columnspan=2, sticky='we', padx=5)

    # Entry cho cân nặng
   can_nang_tu_entry = tk.Entry(Loc)
   can_nang_den_entry = tk.Entry(Loc)
   can_nang_tu_entry.grid(row=4, column=1, sticky='we', padx=5)
   tk.Label(Loc, text="đến", bg="lightblue").grid(row=4, column=2)
   can_nang_den_entry.grid(row=4, column=3, sticky='we', padx=5)

   # Combobox thời gian (Ngày - Tháng - Năm)
   ngay_cb = ttk.Combobox(Loc, values=[""] + [str(i).zfill(2) for i in range(1, 32)], width=5, state="readonly")
   thang_cb = ttk.Combobox(Loc, values=[""] + [str(i).zfill(2) for i in range(1, 13)], width=5, state="readonly")
   nam_cb = ttk.Combobox(Loc, values=[""] + [str(i) for i in range(2020, 2031)], width=7, state="readonly")
   ngay_cb.grid(row=5, column=1, padx=2)
   thang_cb.grid(row=5, column=2, padx=2)
   nam_cb.grid(row=5, column=3, padx=2)

   def apply_filters():
    # Lấy giá trị từ các widget
    loai_don = loai_don_cb.get()
    loai_sp = loai_san_pham_cb.get()
    ptgh = phuong_thuc_cb.get()
    trang_thai = trang_thai_cb.get()
    can_nang_tu = can_nang_tu_entry.get()
    can_nang_den = can_nang_den_entry.get()
    ngay = ngay_cb.get()
    thang = thang_cb.get()
    nam = nam_cb.get()

    # Kiểm tra nếu tất cả các trường đều rỗng
    if not any([loai_don, loai_sp, ptgh, trang_thai, can_nang_tu, can_nang_den, ngay, thang, nam]):
        messagebox.showwarning("Cảnh báo", "Vui lòng chọn ít nhất một bộ lọc.")
        return

    # Đọc tất cả đơn hàng
    orders = order_manager.FileRead()
    filtered_orders = []

    # Áp dụng các bộ lọc
    for order in orders:
        match = True

        if loai_don:
            if loai_don == "Quốc tế" and not order.Ma.startswith("QT"):
                match = False
            elif loai_don == "Nội địa" and not order.Ma.startswith("VN"):
                match = False

        if loai_sp and match:
            loai_sp_map = {
                "Điện tử": "DT",
                "Gia dụng": "GD",
                "Thể thao": "TT",
                "Thời trang": "TR",
                "Khác": "LT"
            }
            if order.Ma[2:4] != loai_sp_map.get(loai_sp, ""):
                match = False

        if ptgh and match:
            if ptgh == "COD" and order.Ma[4] != 'C':
                match = False
            elif ptgh == "Non-COD" and order.Ma[4] != 'N':
                match = False

        if trang_thai and match:
            if order.status != trang_thai:
                match = False

        if (can_nang_tu or can_nang_den) and match:
            try:
                weight = float(order.cannang)
                if can_nang_tu and weight < float(can_nang_tu):
                    match = False
                if can_nang_den and weight > float(can_nang_den):
                    match = False
            except ValueError:
                match = False

        if (ngay or thang or nam) and match:
            try:
                order_date = datetime.datetime.strptime(order.NgayDatHang, "%d/%m/%Y %H:%M:%S")
                if ngay and order_date.day != int(ngay):
                    match = False
                if thang and order_date.month != int(thang):
                    match = False
                if nam and order_date.year != int(nam):
                    match = False
            except (ValueError, AttributeError):
                match = False

        if match:
            filtered_orders.append(order)

    # Cấu hình tag màu theo trạng thái
    tree.tag_configure('delivered', background='#d4edda')     # Đã giao – xanh lá
    tree.tag_configure('shipping', background='#fff3cd')      # Đang vận chuyển – vàng
    tree.tag_configure('canceled', background='#f8d7da')      # Đã hủy – đỏ nhạt
    tree.tag_configure('in_stock', background='#d1ecf1')      # Đã nhập kho – xanh nhạt
    tree.tag_configure('delivering', background='#cce5ff')    # Đang giao – xanh dương

    # Cập nhật Treeview
    tree.delete(*tree.get_children())
    for index, order in enumerate(filtered_orders, start=1):
        # Gán tag phù hợp theo trạng thái
        tags = ()
        if order.status == "Đã giao":
            tags = ('delivered',)
        elif order.status == "Đang vận chuyển":
            tags = ('shipping',)
        elif order.status == "Đã hủy":
            tags = ('canceled',)
        elif order.status == "Đã nhập kho":
            tags = ('in_stock',)
        elif order.status == "Đang giao":
            tags = ('delivering',)

        tree.insert("", "end", values=(
            index,
            order.Ma,
            order.TenH,
            order.NNhan,
            order.status,
            order.TO,
            order.NgayDatHang
        ), tags=tags)

    messagebox.showinfo("Thông báo", f"Đã tìm thấy {len(filtered_orders)} đơn hàng phù hợp.")
    Loc.destroy()
    # Nút Đồng ý và Hủy
   button_frame = tk.Frame(Loc, bg="lightblue")
   button_frame.grid(row=6, column=0, columnspan=4, pady=15)
   tk.Button(button_frame, text="Đồng ý", command=apply_filters, bg="#3b78d8", fg="white", width=15).pack(side="left", padx=10)
   tk.Button(button_frame, text="Hủy", command=Loc.destroy, bg="#f44336", fg="white", width=15).pack(side="left", padx=10)

#cửa sổ thêm mới đơn hàng
def open_add_order_window(parent_window, tree, order_manager):
    if current_file == None:
        messagebox.showwarning("Thông báo", "Bạn chưa có file để thêm đơn hàng.")
        return
    add_window = tk.Toplevel(parent_window)
    add_window.title("Thêm mới đơn hàng")
    add_window.geometry("650x520")
    add_window.configure(bg="lightblue")
    add_window.iconbitmap(resource_path("ico\Find.ico"))
    fields = ["Loại đơn", "Loại hàng", "Phương thức giao hàng", 
              "Tên đơn hàng", "Tên shop", "Tên người nhận", "Cân nặng (kg)", 
              "Đơn vị vận chuyển", "Trạng thái", "Địa chỉ nhận hàng", "Địa chỉ lấy hàng",
              "SĐT người nhận", "SĐT người gửi", "COD"]
    entries = {}

    # Biến riêng cho các combo cần theo dõi
    ptgh_var = tk.StringVar()
    loaidon_var = tk.StringVar()
    loaihang_var = tk.StringVar()

    for row, field in enumerate(fields):
        tk.Label(add_window, text=f"{field}:", bg="lightblue").grid(row=row, column=0, padx=20, pady=5, sticky="e")

        if field == "Loại đơn":
            combo = ttk.Combobox(add_window, width=27, textvariable=loaidon_var, state="readonly")
            combo['values'] = ("Quốc tế", "Nội địa")
            combo.set("Nội địa")
            combo.grid(row=row, column=1, padx=20, pady=5)
            entries[field] = combo

        elif field == "Loại hàng":
            combo = ttk.Combobox(add_window, width=27, textvariable=loaihang_var, state="readonly")
            combo['values'] = ("Điện tử", "Gia dụng", "Thể thao", "Thời trang", "Khác")
            combo.set("Điện tử")
            combo.grid(row=row, column=1, padx=20, pady=5)
            entries[field] = combo

        elif field == "Phương thức giao hàng":
            combo = ttk.Combobox(add_window, width=27, textvariable=ptgh_var, state="readonly")
            combo['values'] = ("COD", "None-COD")
            combo.set("None-COD")
            combo.grid(row=row, column=1, padx=20, pady=5)
            entries[field] = combo

        elif field == "Trạng thái":
            combo = ttk.Combobox(add_window, width=27, state="readonly")
            combo['values'] = ("Đang giao", "Đã giao", "Đang vận chuyển", "Đã hủy", "Đã nhập kho")
            combo.set("Đang giao")
            combo.grid(row=row, column=1, padx=20, pady=5)
            entries[field] = combo

        elif field == "COD":
            cod_entry = tk.Entry(add_window, width=30)
            cod_entry.insert(0, "0")
            cod_entry.config(state="disabled")
            cod_entry.grid(row=row, column=1, padx=20, pady=5)
            entries[field] = cod_entry

        else:
            entry = tk.Entry(add_window, width=30)
            entry.grid(row=row, column=1, padx=20, pady=5)
            entries[field] = entry

    def toggle_cod_field(event=None):
        cod_field = entries["COD"]
        if ptgh_var.get() == "COD":
            cod_field.config(state="normal")
            cod_field.delete(0, tk.END)
            cod_field.insert(0, "1")
        else:
            cod_field.delete(0, tk.END)
            cod_field.insert(0, "0")
            cod_field.config(state="disabled")

    ptgh_var.trace_add("write", lambda *args: toggle_cod_field())
    toggle_cod_field()

    def save_order():
        # Tạo mã đơn hàng tự động tăng dần
        loai_don = loaidon_var.get()
        loai_hang = loaihang_var.get()
        ptgh = ptgh_var.get()

        prefix_map = {"Quốc tế": "QT", "Nội địa": "VN"}
        loaihang_map = {"Điện tử": "DT", "Gia dụng": "GD", "Thể thao": "TT", "Thời trang": "TR", "Khác": "LT"}
        ptgh_code = "C" if ptgh == "COD" else "N"
        prefix = prefix_map[loai_don] + loaihang_map[loai_hang] + ptgh_code

        existing_orders = order_manager.FileRead()
        existing_suffixes = [
            int(order.Ma[5:]) for order in existing_orders
            if order.Ma.startswith(prefix) and order.Ma[5:].isdigit()
        ]
        next_number = max(existing_suffixes, default=0) + 1
        full_code = f"{prefix}{str(next_number).zfill(5)}"

        # Kiểm tra dữ liệu đầu vào (trừ mã đã tự tạo và COD đã xử lý)
        values = [entries[field].get().strip() for field in fields if field != "COD"]
        if any(not v for v in values):
            messagebox.showwarning("Thiếu thông tin", "Vui lòng nhập đầy đủ thông tin.")
            return
                # Kiểm tra giá trị cân nặng là số thực dương
        try:
            weight = float(entries["Cân nặng (kg)"].get())
            if weight <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Lỗi", "Cân nặng không hợp lệ!.")
            return

        current_date = datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        donhang = DonHang(
            full_code,
            entries["Tên đơn hàng"].get(),
            entries["Tên shop"].get(),
            entries["Tên người nhận"].get(),
            entries["Cân nặng (kg)"].get(),
            entries["Đơn vị vận chuyển"].get(),
            entries["Trạng thái"].get(),
            entries["Địa chỉ nhận hàng"].get(),
            entries["Địa chỉ lấy hàng"].get(),
            entries["SĐT người nhận"].get(),
            entries["SĐT người gửi"].get(),
            entries["COD"].get(),
            current_date
        )

        existing_orders.append(donhang)
        order_manager.save_orders(existing_orders)

        tree.insert("", "end", values=(
            "", donhang.Ma, donhang.TenH, donhang.NNhan,
            donhang.status, donhang.TO, donhang.NgayDatHang
        ))

        messagebox.showinfo("Thành công", f"Đơn hàng {donhang.Ma} đã được thêm mới.")
        add_window.destroy()

    tk.Button(add_window, text="Lưu", bg="lightgreen", command=save_order).grid(row=len(fields)+1, column=0, columnspan=2, pady=15)
#xử lý mã đơn hàng
def parse_order_code(code):
    if not code or len(code) != 10:
        return (3, 0)  # Mã không hợp lệ
    
    # Phân loại ưu tiên: QT > VN > Khác
    loai_don = code[:2]
    if loai_don == "QT":
        priority = 0
    elif loai_don == "VN":
        priority = 1
    else:
        priority = 2
    
    # Phân loại sản phẩm: DT > GD > TT > TR > LT
    loai_sp = code[2:4]
    if loai_sp == "DT":
        sp_priority = 0
    elif loai_sp == "GD":
        sp_priority = 1
    elif loai_sp == "TT":
        sp_priority = 2
    elif loai_sp == "TR":
        sp_priority = 3
    else:  # LT
        sp_priority = 4
    
    # Hình thức giao hàng: COD (C) ưu tiên hơn None-COD (N)
    hinh_thuc = code[4]
    delivery_priority = 0 if hinh_thuc == "C" else 1
    
    # Số thứ tự (5 ký tự cuối)
    try:
        so_thu_tu = int(code[5:])
    except ValueError:
        so_thu_tu = 0
    
    return (priority, sp_priority, delivery_priority, so_thu_tu)
def sort_by_order_code():
    sort_states["Mã đơn hàng"] = not sort_states["Mã đơn hàng"]
    ascending = sort_states["Mã đơn hàng"]

    items = tree.get_children()
    data = []
    for item in items:
        values = tree.item(item)["values"]
        data.append((parse_order_code(values[1]), values))

    data.sort(key=lambda x: x[0], reverse=not ascending)

    update_treeview(data)
# Hàm phân tích ngày đặt hàng
def parse_order_date(date_str):
    try:
        return datetime.datetime.strptime(date_str, "%d/%m/%Y %H:%M:%S")
    except ValueError:
        return datetime.datetime.min  # hoặc datetime.datetime(1900, 1, 1)
# Trạng thái sắp xếp từng cột
sort_states = {
    "Mã đơn hàng": True,
    "Ngày đặt hàng": True,
    "Trạng thái": True 
}
def export_to_excel():
    """
    Xuất toàn bộ thông tin đơn hàng ra file Excel (.xlsx)
    Xử lý đầy đủ các trường hợp lỗi có thể xảy ra
    """
    try:
        # Kiểm tra phụ thuộc
        
        # Kiểm tra file hiện tại
        if not current_file:
            raise ValueError("Chưa chọn file dữ liệu làm việc")

        # Đọc dữ liệu
        try:
            orders = order_manager.FileRead()
            if not orders:
                raise ValueError("Không có đơn hàng nào để xuất")
        except Exception as e:
            raise IOError(f"Lỗi đọc file dữ liệu: {str(e)}")

        # Tạo workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Đơn hàng"[:31]  # Giới hạn 31 ký tự cho tên sheet

        # Định nghĩa cột
        columns = [
            ("STT", 8),
            ("Mã đơn hàng", 15),
            ("Tên đơn hàng", 25),
            ("Tên shop", 20),
            ("Người nhận", 20),
            ("Cân nặng (kg)", 12),
            ("Đơn vị VC", 15),
            ("Trạng thái", 15),
            ("Địa chỉ nhận", 30),
            ("Địa chỉ gửi", 30),
            ("SĐT nhận", 15),
            ("SĐT gửi", 15),
            ("COD", 12),
            ("Ngày đặt", 20),
            ("Ghi chú", 40)
        ]

        # Tạo header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="0072C6",
            end_color="0072C6",
            fill_type="solid"
        )

        for col_idx, (col_name, width) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Đổ dữ liệu
        for row_idx, order in enumerate(orders, 2):
            try:
                order_dict = order.to_dict()
                
                # Định dạng ngày nếu cần
                order_date = order_dict.get("Ngày đặt hàng", "")
                if isinstance(order_date, datetime.datetime):
                    order_date = order_date.strftime("%d/%m/%Y %H:%M")
                
                # Các giá trị cần định dạng đặc biệt
                cod_value = order_dict.get("COD", 0)
                try:
                    cod_value = float(cod_value) if str(cod_value).isdigit() else 0
                except:
                    cod_value = 0

                weight = order_dict.get("Cân nặng (kg)", 0)
                try:
                    weight = float(weight)
                except:
                    weight = 0

                # Ghi dữ liệu
                ws.cell(row=row_idx, column=1, value=row_idx-1).alignment = Alignment(horizontal="center")
                
                data_mapping = [
                    order_dict.get("Mã đơn hàng", ""),
                    order_dict.get("Tên đơn hàng", ""),
                    order_dict.get("Tên shop", ""),
                    order_dict.get("Tên người nhận", ""),
                    weight,
                    order_dict.get("Đơn vị vận chuyển", ""),
                    order_dict.get("Trạng thái", ""),
                    order_dict.get("Địa chỉ nhận hàng", ""),
                    order_dict.get("Địa chỉ lấy hàng", ""),
                    order_dict.get("SĐT người nhận", ""),
                    order_dict.get("SĐT người gửi", ""),
                    cod_value,
                    order_date,
                    ""  # Ghi chú
                ]

                for col_idx, value in enumerate(data_mapping, 2):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    
                    # Định dạng số
                    if col_idx == 6:  # Cân nặng
                        cell.number_format = '0.00'
                    elif col_idx == 13:  # COD
                        cell.number_format = '#,##0'
                    elif col_idx == 14:  # Ngày đặt
                        cell.alignment = Alignment(horizontal="center")

            except Exception as row_error:
                print(f"Lỗi dòng {row_idx}: {str(row_error)}")
                continue

        # Thiết lập bộ lọc
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(orders)+1}"

        # Lưu file
        default_name = f"DonHang_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            title="Lưu file báo cáo đơn hàng",
            initialfile=default_name
        )

        if not file_path:
            return  # Người dùng hủy

        # Thử các phương pháp lưu khác nhau nếu cần
        try:
            wb.save(file_path)
        except PermissionError:
            raise IOError("Không có quyền ghi file. Có thể file đang mở bởi chương trình khác")
        except Exception as save_error:
            raise IOError(f"Lỗi khi lưu file: {str(save_error)}")

        # Thông báo thành công
        success_msg = f"""
        Xuất file thành công!
        - Vị trí: {file_path}
        - Tổng đơn: {len(orders)}
        - Đã giao: {sum(1 for o in orders if o.status == 'Đã giao')}
        - Đang xử lý: {sum(1 for o in orders if o.status != 'Đã giao')}
        """
        messagebox.showinfo("Hoàn thành", success_msg.strip())

    except ImportError as ie:
        messagebox.showerror("Lỗi thư viện", str(ie))
    except ValueError as ve:
        messagebox.showwarning("Cảnh báo", str(ve))
    except IOError as ioe:
        messagebox.showerror("Lỗi IO", f"Lỗi hệ thống file:\n{str(ioe)}")

    except Exception as e:
        traceback.print_exc()  # ✅ In lỗi đầy đủ ra terminal/console
        messagebox.showerror("Lỗi không xác định", f"Có lỗi nghiêm trọng xảy ra:\n{str(e)}")

    finally:
        # Dọn dẹp nếu cần
        if 'wb' in locals():
            try:
                wb.close()
            except:
                pass
#sắp xếp theo trạng thái đơn hàng
def sort_by_status():
    sort_states["Trạng thái"] = not sort_states["Trạng thái"]
    ascending = sort_states["Trạng thái"]

    status_order = {
        "Đã giao": 0,
        "Đang giao": 1,
        "Đang vận chuyển": 2,
        "Đã nhập kho": 3,
        "Đã hủy": 4
    }

    items = tree.get_children()
    data = []

    for item in items:
        values = tree.item(item)["values"]
        status = values[4]  # cột trạng thái
        status_priority = status_order.get(status, 99)
        data.append((status_priority, values))

    data.sort(key=lambda x: x[0], reverse=not ascending)
    update_treeview(data)
# Sắp xếp theo mã đơn hàng
def sort_by_order_code():
    sort_states["Mã đơn hàng"] = not sort_states["Mã đơn hàng"]
    ascending = sort_states["Mã đơn hàng"]

    items = tree.get_children()
    data = []
    for item in items:
        values = tree.item(item)["values"]
        data.append((parse_order_code(values[1]), values))

    data.sort(key=lambda x: x[0], reverse=not ascending)
    update_treeview(data)
#reset giá trị khi người dùng đăng xuất 
# Sắp xếp theo ngày đặt hàng
def sort_by_order_date():
    sort_states["Ngày đặt hàng"] = not sort_states["Ngày đặt hàng"]
    ascending = sort_states["Ngày đặt hàng"]

    items = tree.get_children()
    data = []
    for item in items:
        values = tree.item(item)["values"]
        date = parse_order_date(values[6])
        data.append((date, values))

    data.sort(key=lambda x: x[0], reverse=not ascending)

    update_treeview(data)
# Cập nhật lại treeview với dữ liệu đã sắp xếp
def update_treeview(data):
    tree.delete(*tree.get_children())

    # Thiết lập lại màu theo trạng thái
    tree.tag_configure('delivered', background='#d4edda')
    tree.tag_configure('shipping', background='#fff3cd')
    tree.tag_configure('canceled', background='#f8d7da')
    tree.tag_configure('in_stock', background='#d1ecf1')
    tree.tag_configure('delivering', background='#cce5ff')

    for index, (_, values) in enumerate(data, 1):
        values = list(values)
        values[0] = index  # cập nhật lại STT

        # Xác định tag theo trạng thái
        status = values[4]
        tags = ()
        if status == "Đã giao":
            tags = ('delivered',)
        elif status == "Đang vận chuyển":
            tags = ('shipping',)
        elif status == "Đã hủy":
            tags = ('canceled',)
        elif status == "Đã nhập kho":
            tags = ('in_stock',)
        elif status == "Đang giao":
            tags = ('delivering',)

        tree.insert("", "end", values=values, tags=tags)
#chỉnh sửa đơn hàng được chọn
def edit_selected_order():
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Lỗi", "Vui lòng chọn một đơn hàng để chỉnh sửa!")
        return

    values = tree.item(selected_item, "values")
    old_order_id = values[1]

    orders = order_manager.FileRead()
    original_order = None
    
    for order in orders:
        if order.Ma == old_order_id:
            original_order = order
            break

    if not original_order:
        messagebox.showerror("Lỗi", "Không tìm thấy đơn hàng để chỉnh sửa!")
        return

    edit_window = tk.Toplevel()
    edit_window.title("Chỉnh sửa đơn hàng")
    edit_window.geometry("750x500")
    edit_window.configure(bg="lightblue")
    edit_window.iconbitmap(resource_path("ico\edit.ico"))
    frame = tk.Frame(edit_window, bg="lightblue")
    frame.place(relx=0.5, rely=0.5, anchor="center")

    fields = [
        "Mã đơn hàng", "Tên đơn hàng", "Tên shop", "Tên người nhận", 
        "Cân nặng (kg)", "Đơn vị vận chuyển", "Trạng thái",
        "Địa chỉ nhận hàng", "Địa chỉ lấy hàng", 
        "SĐT người nhận", "SĐT người gửi", "COD"
    ]
    
    entries = {}

    # Kiểm tra ký tự thứ 5 của mã đơn hàng
    ma_don = original_order.Ma
    char5 = ma_don[4] if len(ma_don) >= 5 else ""
    is_N = char5 == "N"
    is_C = char5 == "C"

    for row, field in enumerate(fields):
        tk.Label(frame, text=f"{field}:", bg="lightblue", font=("Arial", 10, "bold")).grid(
            row=row, column=0, padx=10, pady=5, sticky="e")
        
        if field == "Trạng thái":
            status_combobox = ttk.Combobox(
                frame,
                width=37,
                values=["Đang giao", "Đã giao", "Đang vận chuyển", "Đã hủy", "Đã nhập kho"],
                state="readonly"
            )
            status_combobox.set(original_order.status)
            status_combobox.grid(row=row, column=1, padx=10, pady=5)
            entries[field] = status_combobox
        elif field == "COD":
            cod_entry = tk.Entry(frame, width=40)
            cod_value = "0" if is_N else original_order.COD
            cod_entry.insert(0, cod_value)
            if is_N:
                cod_entry.config(state="disabled")
            cod_entry.grid(row=row, column=1, padx=10, pady=5)
            entries[field] = cod_entry
        else:
            entry = tk.Entry(frame, width=40)
            entry.grid(row=row, column=1, padx=10, pady=5)
            entries[field] = entry

    # Điền thông tin vào các trường
    entries["Mã đơn hàng"].insert(0, original_order.Ma)
    entries["Tên đơn hàng"].insert(0, original_order.TenH)
    entries["Tên shop"].insert(0, original_order.NguoiGui)
    entries["Tên người nhận"].insert(0, original_order.NNhan)
    entries["Cân nặng (kg)"].insert(0, original_order.cannang)
    entries["Đơn vị vận chuyển"].insert(0, original_order.DVVC)
    entries["Địa chỉ nhận hàng"].insert(0, original_order.TO)
    entries["Địa chỉ lấy hàng"].insert(0, original_order.From)
    entries["SĐT người nhận"].insert(0, original_order.Pnumber1)
    entries["SĐT người gửi"].insert(0, original_order.Pnumber2)
    entries["Mã đơn hàng"].config(state='disabled')

    # Hiển thị ngày đặt hàng
    tk.Label(frame, text="Ngày đặt hàng:", bg="lightblue", font=("Arial", 10, "bold")).grid(
        row=len(fields), column=0, padx=10, pady=5, sticky="e")
    date_label = tk.Label(frame, text=original_order.NgayDatHang, bg="lightblue")
    date_label.grid(row=len(fields), column=1, padx=10, pady=5, sticky="w")

    def save_edited_order():
        edited_values = {}
        for key in fields:
            edited_values[key] = entries[key].get()

        # Kiểm tra dữ liệu (trừ COD nếu bị khóa)
        for key, value in edited_values.items():
            if not value.strip() and key != "COD":
                messagebox.showwarning("Lỗi", f"Vui lòng nhập đầy đủ thông tin cho trường {key}!")
                return

        # Cập nhật thông tin
        original_order.TenH = edited_values["Tên đơn hàng"]
        original_order.NguoiGui = edited_values["Tên shop"]
        original_order.NNhan = edited_values["Tên người nhận"]
        original_order.cannang = edited_values["Cân nặng (kg)"]
        original_order.DVVC = edited_values["Đơn vị vận chuyển"]
        original_order.status = edited_values["Trạng thái"]
        original_order.TO = edited_values["Địa chỉ nhận hàng"]
        original_order.From = edited_values["Địa chỉ lấy hàng"]
        original_order.Pnumber1 = edited_values["SĐT người nhận"]
        original_order.Pnumber2 = edited_values["SĐT người gửi"]
        original_order.COD = "0" if is_N else edited_values["COD"]

        order_manager.save_orders(orders)
        load_orders()
        
        messagebox.showinfo("Thành công", "Đơn hàng đã được cập nhật!")
        edit_window.destroy()

    btn_save = tk.Button(frame, text="Lưu thay đổi", bg="lightgreen", font=("Arial", 10, "bold"),
                        command=save_edited_order)
    btn_save.grid(row=len(fields)+1, column=0, columnspan=2, pady=10)    
#cập nhập trạng thái
def update_stats():
    orders = order_manager.FileRead()

    total_orders = len(orders)
    delivered_count = sum(1 for order in orders if order.status == "Đã giao")
    delivering_count = sum(1 for order in orders if order.status == "Đang giao")
    shipping_count = sum(1 for order in orders if order.status == "Đang vận chuyển")
    canceled_count = sum(1 for order in orders if order.status == "Đã hủy")
    Kho = sum(1 for order in orders if order.status == "Đã nhập kho")

    total_label.config(text=f"Tổng đơn hàng: {total_orders}")
    delivered_label.config(text=f"Đã giao: {delivered_count}")
    delivering_label.config(text=f"Đang giao: {delivering_count}")
    Kho_van.config(text=f"Trong kho:{Kho}")
    shipping_label.config(text=f"Đang vận chuyển: {shipping_count}")
    canceled_label.config(text=f"Đã hủy: {canceled_count}")
#tải đơn hàng
def load_orders():
    global order_manager, current_file, tree
    if current_file == None: messagebox.showinfo("Thông báo","Vui lòng chọn file để làm việc")
    else:
        order_manager = QuanLyDonHang(current_file)
        # Cập nhật Treeview
        tree.delete(*tree.get_children())
        orders = order_manager.FileRead()
        
        # Thiết lập màu sắc theo trạng thái
        tree.tag_configure('delivered', background='#d4edda')
        tree.tag_configure('shipping', background='#fff3cd')
        tree.tag_configure('canceled', background='#f8d7da')
        tree.tag_configure('in_stock', background='#d1ecf1')
        tree.tag_configure('delivering', background='#cce5ff')
        
        for index, order in enumerate(orders, start=1):
            tags = ()
            if order.status == "Đã giao":
                tags = ('delivered',)
            elif order.status == "Đang vận chuyển":
                tags = ('shipping',)
            elif order.status == "Đã hủy":
                tags = ('canceled',)
            elif order.status == "Đã nhập kho":
                tags = ('in_stock',)
            elif order.status == "Đang giao":
                tags = ('delivering',)
                
            tree.insert("", "end", values=(
                index, 
                order.Ma, 
                order.TenH, 
                order.NNhan, 
                order.status, 
                order.TO,
                order.NgayDatHang
            ), tags=tags)
        
        update_stats()
#tìm kiếm đơn hàng
def search_orders(entry):
    keyword = entry.get().strip().lower()
    if not keyword or keyword == "nhập tên, mã đơn hàng để tìm kiếm":
        load_orders()
        return
    orders = order_manager.FileRead()
    filtered_orders = [
        order for order in orders
        if keyword in order.Ma.lower() or keyword in order.TenH.lower()
    ]

    tree.delete(*tree.get_children())  # Xóa dữ liệu cũ
    for index, order in enumerate(filtered_orders, start=1):
        status = order.status
        tag = ()
        if status == "Đã giao":
            tag = ('delivered',)
        elif status == "Đang vận chuyển":
            tag = ('shipping',)
        elif status == "Đã hủy":
            tag = ('canceled',)
        elif status == "Đã nhập kho":
            tag = ('in_stock',)
        elif status == "Đang giao":
            tag = ('delivering',)

        tree.insert("", "end", values=(
            index,
            order.Ma,
            order.TenH,
            order.NNhan,
            order.status,
            order.TO,
            order.NgayDatHang
        ), tags=tag)

    update_stats()
tree = None
total_label = None
delivered_label = None
delivering_label = None
shipping_label = None
canceled_label = None
Kho_van = None   
global current_file
current_file = None
order_manager = QuanLyDonHang(current_file)
#Các hàm xư lý trong user_file_management
def load_user_files(tree):
    """Tải danh sách file của user kèm thông tin chi tiết"""
    tree.delete(*tree.get_children())
    
    users = load_users()
    current_user = next((u for u in users["users"] if u["username"] == current_username), None)
    
    if not current_user or not current_user.get("data"):
        tree.insert("", "end", values=("1", "Không có file nào", "", "", ""))
        return
    
    for idx, filepath in enumerate(current_user["data"], 1):
        try:
            # Lấy thông tin file
            size_kb = os.path.getsize(filepath) / 1024
            manager = QuanLyDonHang(filepath)
            orders = manager.FileRead()
            num_orders = len(orders)
            
            # Thêm vào treeview
            tree.insert(
                "", 
                "end", 
                values=(
                    idx,
                    os.path.basename(filepath),
                    filepath,
                    f"{size_kb:.2f} KB",
                    num_orders
                ),
                tags=("valid",)
            )
        except Exception as e:
            tree.insert(
                "", 
                "end", 
                values=(
                    idx,
                    os.path.basename(filepath),
                    filepath,
                    "Lỗi",
                    "N/A"
                ),
                tags=("error",)
            )
    
    # Định dạng màu
    tree.tag_configure("valid", background="#ffffff")
    tree.tag_configure("error", background="#ffdddd")
def add_user_file(tree):
    """Mở hộp thoại chọn file và thêm vào danh sách"""
    filepath = filedialog.askopenfilename(
        title="Chọn file JSON",
        filetypes=[("JSON files", "*.json")]
    )
    
    if not filepath:
        return

    users = load_users()
    for user in users["users"]:
        if user["username"] == current_username:
            if "data" not in user:
                user["data"] = []
            
            # Kiểm tra trùng lặp
            if filepath in user["data"]:
                messagebox.showwarning("Cảnh báo", "File đã tồn tại trong danh sách!")
                return
                
            user["data"].append(filepath)
            break
    
    # Lưu lại và cập nhật giao diện
    with open("User.json", "w", encoding="utf-8") as f:
        json.dump(users, f, indent=4)
    
    load_user_files(tree)
    messagebox.showinfo("Thành công", f"Đã thêm file: {os.path.basename(filepath)}")
def delete_user_file(tree):
    """Xóa file được chọn khỏi danh sách, nếu là file đơn hàng của người dùng thì làm rỗng nội dung"""

    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Cảnh báo", "Vui lòng chọn file cần xóa!")
        return

    filepath = tree.item(selected_item[0])["values"][2]  # Lấy đường dẫn từ cột 3
    filename = os.path.basename(filepath)

    users = load_users()
    for user in users["users"]:
        if user["username"] == current_username and "data" in user:
            if filepath in user["data"]:
                user["data"].remove(filepath)
                break

    # Nếu là file orders của người dùng → ghi rỗng file
    if filename == f"{current_username}_orders.json":
        try:
            with open(filepath, "w", encoding="utf-8") as f:
                json.dump([], f, indent=4)  # hoặc {} nếu file chứa dict thay vì list
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể làm rỗng file: {e}")
            return

    # Lưu lại danh sách người dùng đã cập nhật
    with open("User.json", "w", encoding="utf-8") as f:
        json.dump(users, f, indent=4)

    # Cập nhật giao diện
    load_user_files(tree)
    
    global current_file
    current_file = None
    messagebox.showinfo("Thành công", "Đã xóa file khỏi danh sách!")

def load_selected_file(tree):
    """Tải file được chọn vào chương trình chính"""
    selected_item = tree.selection()
    if not selected_item:
        messagebox.showwarning("Cảnh báo", "Vui lòng chọn file cần tải!")
        return
    
    filepath = tree.item(selected_item[0])["values"][2]
    
    try:
        global order_manager, current_file
        current_file = filepath
        order_manager = QuanLyDonHang(current_file)
        load_orders()  # Hàm tải đơn hàng vào Treeview chính
        messagebox.showinfo("Thành công", f"Đã tải file: {os.path.basename(filepath)}")
        QL.destroy()
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể tải file!\n{str(e)}")
#Trợ giúp
def help_window():
    help_win = tk.Toplevel()
    help_win.title("Trợ giúp")
    help_win.geometry("650x550")
    help_win.iconbitmap(resource_path("ico\OIP.ico")) 
    help_win.configure(bg="#f0f0f0")  # Nền xám nhạt
    # Frame chứa Text và Scrollbar
    frame = tk.Frame(help_win, bg="#f0f0f0")
    frame.pack(fill="both", expand=True, padx=10, pady=10)
    # Scrollbar
    scrollbar = tk.Scrollbar(frame)
    scrollbar.pack(side="right", fill="y")
    # Text widget
    text_widget = tk.Text(
        frame,
        wrap="word",
        yscrollcommand=scrollbar.set,
        font=("Segoe UI", 11),
        bg="#f5f5f5",            # Nền xám sáng
        fg="#202020",            # Màu chữ gần đen
        insertbackground="black",  # Con trỏ chữ đen
        relief="flat",
        padx=10,
        pady=10
    )
    text_widget.pack(fill="both", expand=True)
    scrollbar.config(command=text_widget.yview)
    help_text = """\
    HƯỚNG DẪN SỬ DỤNG: 
    [Đăng nhập]:
        - Để đăng nhập vào hệ thống, bạn cần nhập tên người dùng và mật khẩu
        - Nếu bạn là người dùng bình thường, hãy nhập tên người dùng và mật khẩu của bạn.
        - Nếu bạn quên mật khẩu, hãy liên hệ với quản trị viên để được hỗ trợ.
        - Vui lòng truy cập "Menu -> Quản lý" để chọn file dữ liệu và bắt đầu làm việc
    [Đăng ký]:
        Vui lòng Nhập đầy đủ thông tin để tiến hành đăng kí tài khoản
        Quản trị viên có thể tạo tài khoản cho người dùng mới bằng cách chọn "Menu - >Quản lý -> Thên mới" từ menu.
    [DỮ LIỆU]:
        - Để tải dữ liệu đơn hàng, bạn cần chọn file JSON chứa dữ liệu đơn hàng.
        - Dữ liệu đơn hàng phải được định dạng đúng theo yêu cầu của hệ thống.
        - Nếu bạn không có file dữ liệu, hãy liên hệ với quản trị viên để được hỗ trợ.
         [ĐỊNH DẠNG DỮ LIỆU ĐƠN HÀNG]:
         {
           [    "Mã đơn hàng": ,
                "Tên đơn hàng": ,
                "Tên shop": ,
                "Tên người nhận": ,
                "Cân nặng (kg)": ,
                "Đơn vị vận chuyển": ,
                "Trạng thái": ,
                "Địa chỉ nhận hàng": ,
                "Địa chỉ lấy hàng": ,
                "SĐT người nhận": ,
                "SĐT người gửi": ,
                "COD": ,
                "Ngày đặt hàng": 
                ]
            }
    [Các thành phần trong menu]:
        Xuất file xlsx: Xuất danh sách đơn hàng ra file Excel
            #LƯU Ý CHỨC NĂNG NÀY CHỈ DÀNH CHO QUẢN TRỊ VIÊN !
        Quản lý:
            - Để quản lý file dữ liệu và người dùng, chọn "Menu-Quản lý" từ menu.
            - Để xem thông tin chi tiết về người dùng, chọn người dùng và nhấn "Chi tiết người dùng".
            -Người dùng khi chọn tải tất cả sẽ tạo ra file mới với định dạng [tên người dùng]_allDH.json
             ~>Quản trị viên sẽ không thể thao tác thêm chức năng này khi vào giao diện quản lý người dùng được chọn
            - Chọn thêm, xóa, tải file để thao tác với file dữ liệu
                    #LƯU Ý FILE PHẢI ĐÚNG ĐỊNH DẠNG MỚI CÓ THỂ ĐỌC ĐƯỢC!
        Để đăng xuất vui lòng chọn "Menu=>Đăng xuất" từ menu.
        Để thoát chương trình vui lòng chọn "Menu=>Thoát" từ menu.
    [CÁC THÀNH PHẦN TẠI GIAO DIỆN CHÍNH]:
        1.Chọn thống kê để thống kê dữ liệu đơn hàng hiện tại
        2.Chọn tìm kiếm để tìm kiếm đơn hàng theo tên hoặc mã đơn hàng
        3.Nhấp vào các header của bảng thông tin để sắp xếp đơn hàng theo trạng thái, mã đơn hàng hoặc ngày đặt hàng
        4.Chọn xem chi tiết để xem chi tiết đơn hàng
        5.Chọn thêm đơn hàng để thêm đơn hàng mới vào danh sách
        6. Chọn xóa đơn hàng để xóa đơn hàng đã chọn
        7. Chọn "Cập nhập lại" để tải lại danh sách đơn hàng
        8. Chọn "Lọc" để áp dụng bộ lọc đơn hàng
    [THÔNG TIN LIÊN HỆ:]
    Nếu bạn cần hỗ trợ hoặc có câu hỏi, vui lòng liên hệ với quản trị viên ứng dụng qua các cách sau:
    Gửi email về hộp thư: 2001231059@hufi.edu.vn
    Gọi điện thoại, nhắn tin đến số: 0353686501
    Liên hệ đến sinh viên: 
    2001230324 - Châu Gia Hưng - khoa Công nghệ thông tin, Trường Đại học Công Thương TP.HCM
    2001230469 - Châu Phát Lộc - khoa Công nghệ thông tin, Trường Đại học Công Thương TP.HCM
    2001231059 - Lô Hoàng Vũ - khoa Công nghệ thông tin, Trường Đại học Công Thương TP.HCM
    Xin cảm ơn bạn đã sử dụng ứng dụng quản lý đơn hàng của chúng tôi!  
    """
    text_widget.insert("1.0", help_text)
    text_widget.config(state="disabled")  # Chỉ đọc
#hiển thị cửa sổ user cho admin
def open_user_detail(tree):
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Thông báo", "Vui lòng chọn người dùng!")
        return

    username = tree.item(selected[0])["values"][0]

    # Tìm user
    users = load_users()["users"]
    user = next((u for u in users if u["username"] == username), None)
    if not user:
        messagebox.showerror("Lỗi", "Không tìm thấy người dùng.")
        return

    QL = tk.Toplevel()
    QL.title(f"Quản lý dữ liệu của: {username}")
    QL.geometry("1000x600")
    QL.iconbitmap(resource_path("ico\KC.ico"))
    tk.Label(QL, text=f"Quản lý file của người dùng: {username}", font=("Arial", 14, "bold")).pack(pady=10)

    frame = tk.Frame(QL)
    frame.pack(fill="both", expand=True, padx=10, pady=10)

    scrollbar = ttk.Scrollbar(frame, orient="vertical")
    scrollbar.pack(side="right", fill="y")

    file_tree = ttk.Treeview(
        frame,
        columns=("STT", "Tên File", "Đường Dẫn", "Kích Thước", "Số Đơn"),
        show="headings",
        yscrollcommand=scrollbar.set
    )
    scrollbar.config(command=file_tree.yview)

    # Cấu hình cột
    file_tree.heading("STT", text="STT")
    file_tree.heading("Tên File", text="Tên File")
    file_tree.heading("Đường Dẫn", text="Đường Dẫn")
    file_tree.heading("Kích Thước", text="Kích Thước (KB)")
    file_tree.heading("Số Đơn", text="Số Đơn")

    file_tree.column("STT", width=50, anchor="center")
    file_tree.column("Tên File", width=150)
    file_tree.column("Đường Dẫn", width=400)
    file_tree.column("Kích Thước", width=120, anchor="center")
    file_tree.column("Số Đơn", width=80, anchor="center")

    file_tree.pack(fill="both", expand=True)

    def load_user_files():
        file_tree.delete(*file_tree.get_children())
        for idx, filepath in enumerate(user.get("data", []), 1):
            try:
                size_kb = os.path.getsize(filepath) / 1024
                manager = QuanLyDonHang(filepath)
                num_orders = len(manager.FileRead())
                file_tree.insert("", "end", values=(idx, os.path.basename(filepath), filepath, f"{size_kb:.2f}", num_orders))
            except:
                file_tree.insert("", "end", values=(idx, "Lỗi file", filepath, "N/A", "N/A"))

    def add_user_file():
        filename = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON files", "*.json")])
        if filename:
            with open(filename, "w", encoding="utf-8") as f:
                json.dump([], f)
            user["data"].append(filename)
            save_user({"users": users})
            load_user_files()
            messagebox.showinfo("Thành công", f"Đã thêm file cho {username}")

    def delete_user_file():
        selected = file_tree.selection()
        if not selected:
            return
        file_path = file_tree.item(selected[0])["values"][2]
        if messagebox.askyesno("Xác nhận", f"Xóa file {file_path} khỏi danh sách người dùng {username}?"):
            if file_path in user["data"]:
                user["data"].remove(file_path)
                save_user({"users": users})
                load_user_files()

    def load_selected_file():
        global current_file
        selected = file_tree.selection()
        if not selected:
            return
        file_path = file_tree.item(selected[0])["values"][2]
        if not os.path.exists(file_path):
            messagebox.showerror("Lỗi", "File không tồn tại.")
            return
        current_file = file_path
        load_orders()
        messagebox.showinfo("Thông báo", f"Đã tải file {os.path.basename(file_path)}")

    def load_all_user_files():
        count = 0
        for file in user.get("data", []):
            try:
                count += len(QuanLyDonHang(file).FileRead())
            except:
                continue
        messagebox.showinfo("Tổng đơn hàng", f"{username} có {count} đơn hàng trong tất cả file.")

    btn_frame = tk.Frame(QL)
    btn_frame.pack(pady=10)

    tk.Button(btn_frame, text="Thêm File", bg="#4CAF50", fg="white", command=add_user_file).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Xóa File", bg="#f44336", fg="white", command=delete_user_file).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Tải File", bg="#2196F3", fg="white", command=load_selected_file).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Tải Tất Cả", bg="#FF9800", fg="white", command=load_all_user_files).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Đóng", bg="#9E9E9E", fg="white", command=QL.destroy).pack(side="left", padx=5)

    load_user_files()
#Xóa dữ liệu đăng nhập
def Reset_giatri():
    current_user_role = None
    current_file = None
    current_username = None
#xóa user
def delete_user(tree):
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Thông báo", "Vui lòng chọn người dùng để xóa.")
        return

    username = tree.item(selected[0])["values"][0]

    if username == "admin":
        messagebox.showerror("Lỗi", "Không thể xóa tài khoản quản trị viên mặc định.")
        return

    confirm = messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa người dùng '{username}'?")
    if not confirm:
        return

    users = load_users()
    users["users"] = [u for u in users["users"] if u["username"] != username]

    with open("User.json", "w", encoding="utf-8") as f:
        json.dump(users, f, indent=4, ensure_ascii=False)

    tree.delete(selected[0])
    messagebox.showinfo("Thành công", f"Đã xóa người dùng '{username}'.")
    AD.destroy()
def Lammoi_user(tree):
    tree.delete(*tree.get_children())  # Xóa toàn bộ dòng hiện tại

    users = load_users()
    for user in users.get("users", []):
        tree.insert("", "end", values=(
            user["username"],
            user["role"],
            len(user.get("data", []))
        ))
def show_user_info(user_data, save_callback=None):
    if user_data.get("username") == "admin":
        messagebox.showwarning("Cảnh báo", "Bạn không thể thao tác lên tài khoản này.")
        return
    info_win = tk.Toplevel()
    info_win.title("Thông tin người dùng")
    info_win.geometry("400x360")
    info_win.configure(bg="#e0f7fa")
    info_win.iconbitmap(resource_path("ico\hehe.ico"))

    tk.Label(info_win, text="Thông tin người dùng", font=("Arial", 14, "bold"), bg="#e0f7fa").pack(pady=10)

    form_frame = tk.Frame(info_win, bg="#e0f7fa")
    form_frame.pack(pady=10)

    # Tên người dùng
    tk.Label(form_frame, text="Tên người dùng:", bg="#e0f7fa", anchor="w", width=15).grid(row=0, column=0, sticky="w")
    tk.Label(form_frame, text=user_data.get("username", ""), bg="#e0f7fa", anchor="w").grid(row=0, column=1, sticky="w")

    # Ngày tạo
    tk.Label(form_frame, text="Ngày tạo:", bg="#e0f7fa", anchor="w", width=15).grid(row=1, column=0, sticky="w")
    tk.Label(form_frame, text=user_data.get("created_at", "N/A"), bg="#e0f7fa", anchor="w").grid(row=1, column=1, sticky="w")

    # Trạng thái
    tk.Label(form_frame, text="Trạng thái:", bg="#e0f7fa", anchor="w", width=15).grid(row=2, column=0, sticky="w")
    status_var = tk.StringVar(value=user_data.get("status", "unknown"))
    status_cb = ttk.Combobox(form_frame, textvariable=status_var, state="readonly", width=15)
    status_cb['values'] = ("active", "disabled", "locked")
    status_cb.grid(row=2, column=1, sticky="w")

    # Vai trò (NEW)
    tk.Label(form_frame, text="Vai trò:", bg="#e0f7fa", anchor="w", width=15).grid(row=3, column=0, sticky="w")
    role_var = tk.StringVar(value=user_data.get("role", "user"))
    role_cb = ttk.Combobox(form_frame, textvariable=role_var, state="readonly", width=15)
    role_cb['values'] = ("user", "admin")
    role_cb.grid(row=3, column=1, sticky="w")

    # Số file dữ liệu
    file_count = len(user_data.get("data", []))
    tk.Label(form_frame, text="Số file dữ liệu:", bg="#e0f7fa", anchor="w", width=15).grid(row=4, column=0, sticky="w")

    file_frame = tk.Frame(form_frame, bg="#e0f7fa")
    file_frame.grid(row=4, column=1, sticky="w")
    tk.Label(file_frame, text=f"{file_count}", bg="#e0f7fa").pack(side="left")

    def show_file_detail():
        files = user_data.get("data", [])
        messagebox.showinfo("Chi tiết file", "\n".join(files) if files else "Không có file nào.")

    detail_btn = tk.Button(
        file_frame,
        text="Chi tiết",
        fg="blue",
        bg="#e0f7fa",
        bd=0,
        cursor="hand2",
        font=("Arial", 9, "underline"),
        command=show_file_detail
    )
    detail_btn.pack(side="left", padx=5)

    # Nút lưu
    def save_status():
        new_status = status_var.get()
        new_role = role_var.get()
        user_data["status"] = new_status
        user_data["role"] = new_role

        users = load_users()
        for u in users["users"]:
            if u["username"] == user_data["username"]:
                u["status"] = new_status
                u["role"] = new_role

        with open("User.json", "w", encoding="utf-8") as f:
            json.dump(users, f, indent=4, ensure_ascii=False)

        if save_callback:
            save_callback()

        messagebox.showinfo("Thành công", f"Đã cập nhật thông tin người dùng '{user_data['username']}'.")

    # Nút Lưu
    tk.Button(info_win, text="Lưu", command=save_status, bg="#4CAF50", fg="white", width=15).pack(pady=15)
    tk.Button(info_win, text="Đóng", command=info_win.destroy, bg="#9E9E9E", fg="black", width=15).pack(pady=5)

#Các Hàm Quản lý
def admin_user_management():
    global AD
    AD = tk.Toplevel()
    AD.title("Quản lý Dữ liệu")
    AD.geometry("900x550")
    AD.configure(bg="#f0f0f0")
    AD.iconbitmap(resource_path("ico\hehe.ico")) 

    # Tiêu đề
    tk.Label(AD, text="Danh sách người dùng", font=("Arial", 13, "bold"), bg="#f0f0f0").pack(pady=10)

    # Treeview hiển thị người dùng
    user_tree = ttk.Treeview(
        AD,
        columns=("Username", "Role", "Files", "Status", "Created"),
        show="headings"
    )

    user_tree.heading("Username", text="Tên người dùng")
    user_tree.heading("Role", text="Vai trò")
    user_tree.heading("Files", text="Số file dữ liệu")
    user_tree.heading("Status", text="Trạng thái")
    user_tree.heading("Created", text="Ngày tạo")

    user_tree.column("Username", width=150)
    user_tree.column("Role", width=80, anchor="center")
    user_tree.column("Files", width=100, anchor="center")
    user_tree.column("Status", width=100, anchor="center")
    user_tree.column("Created", width=180, anchor="center")

    user_tree.pack(fill="both", expand=True, padx=10, pady=5)

    # Hàm tải lại danh sách
    def refresh_user_list(tree_widget):
        tree_widget.delete(*tree_widget.get_children())
        users = load_users()
        for user in users.get("users", []):
            tree_widget.insert("", "end", values=(
                user["username"],
                user["role"],
                len(user.get("data", [])),
                user.get("status", "unknown"),
                user.get("created_at", "N/A")
            ))

    refresh_user_list(user_tree)

    # Hàm lấy user đang chọn
    def get_selected_user(tree_widget):
        selected = tree_widget.selection()
        if not selected:
            messagebox.showwarning("Chưa chọn", "Vui lòng chọn người dùng.")
            return None
        username = tree_widget.item(selected[0])["values"][0]
        users = load_users()["users"]
        return next((u for u in users if u["username"] == username), None)

    # Các nút chức năng
    btn_frame = tk.Frame(AD, bg="#f0f0f0")
    btn_frame.pack(pady=10)

    tk.Button(btn_frame, text="Thêm người dùng", bg="#4CAF50", fg="white",
              command=open_register_window).pack(side="left", padx=5)

    tk.Button(btn_frame, text="Xóa người dùng", bg="#f44336", fg="white",
              command=lambda: delete_user(user_tree)).pack(side="left", padx=5)

    tk.Button(btn_frame, text="Chi tiết người dùng", bg="#2196F3", fg="white",
              command=lambda: open_user_detail(user_tree)).pack(side="left", padx=5)

    tk.Button(btn_frame, text="Xem thông tin", bg="#3F51B5", fg="white",
          command=lambda: (
              (selected := get_selected_user(user_tree)) and show_user_info(selected, save_callback=lambda: refresh_user_list(user_tree))
          ,AD.destroy())).pack(side="left", padx=5)


    tk.Button(btn_frame, text="Cập nhật", bg="#9E9E9E", fg="black",
              command=lambda: refresh_user_list(user_tree)).pack(side="left", padx=5)

    tk.Button(btn_frame, text="Đóng", bg="#9E9E9E", fg="black",
              command=AD.destroy).pack(side="left", padx=5)
def user_file_management():
    global QL
    QL = tk.Toplevel()
    QL.title("Quản lý Dữ liệu Cá nhân")
    QL.geometry("1000x600")
    QL.iconbitmap(resource_path("ico\OIP.ico")) 
    
    tk.Label(QL, text=f"Quản lý file cá nhân",
             font=("Arial", 14, "bold")).pack(pady=10)

    frame = tk.Frame(QL)
    frame.pack(fill="both", expand=True, padx=10, pady=10)

    # Scrollbar
    scrollbar = ttk.Scrollbar(frame, orient="vertical")
    scrollbar.pack(side="right", fill="y")

    # Treeview
    file_tree = ttk.Treeview(
        frame,
        columns=("STT", "Tên File", "Đường Dẫn", "Kích Thước", "Số Đơn"),
        show="headings",
        yscrollcommand=scrollbar.set
    )
    scrollbar.config(command=file_tree.yview)

    # Cấu hình cột
    file_tree.heading("STT", text="STT")
    file_tree.heading("Tên File", text="Tên File")
    file_tree.heading("Đường Dẫn", text="Đường Dẫn")
    file_tree.heading("Kích Thước", text="Kích Thước (KB)")
    file_tree.heading("Số Đơn", text="Số Đơn")

    file_tree.column("STT", width=50, anchor="center")
    file_tree.column("Tên File", width=150)
    file_tree.column("Đường Dẫn", width=400)
    file_tree.column("Kích Thước", width=120, anchor="center")
    file_tree.column("Số Đơn", width=80, anchor="center")

    file_tree.pack(fill="both", expand=True)

    # Load dữ liệu file của user hiện tại
    def load_user_files():
        file_tree.delete(*file_tree.get_children())

        users = load_users()
        current_user = next((u for u in users["users"] if u["username"] == current_username), None)

        if not current_user or not current_user.get("data"):
            return

        for idx, filepath in enumerate(current_user["data"], 1):
            try:
                size_kb = os.path.getsize(filepath) / 1024
                manager = QuanLyDonHang(filepath)
                num_orders = len(manager.FileRead())

                file_tree.insert("", "end", values=(
                    idx,
                    os.path.basename(filepath),
                    filepath,
                    f"{size_kb:.2f}",
                    num_orders
                ))
            except:
                file_tree.insert("", "end", values=(idx, "Lỗi file", filepath, "N/A", "N/A"))

    # Nút thao tác
    btn_frame = tk.Frame(QL)
    btn_frame.pack(pady=10)
    tk.Button(btn_frame, text="Tạo mới", bg="#4682B4", fg="white",command=lambda:new_user_file(file_tree)).pack(padx=5,side="left")
    tk.Button(btn_frame, text="Thêm File", bg="#4CAF50", fg="white",
              command=lambda: add_user_file(file_tree)).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Xóa File", bg="#f44336", fg="white",
              command=lambda: delete_user_file(file_tree)).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Tải File", bg="#2196F3", fg="white",
              command=lambda: load_selected_file(file_tree)).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Tải Tất Cả", bg="#FF9800", fg="white",
              command=lambda:load_all_user_files(file_tree)).pack(side="left", padx=5)
    tk.Button(btn_frame, text="Đóng", bg="#9E9E9E", fg="white",
              command=lambda: QL.destroy()).pack(side="left", padx=5)

    # Tải danh sách file ngay khi mở
    load_user_files() 
#Đổi mật khẩu
def change_password():
    window = tk.Toplevel()
    window.title("Đổi mật khẩu")
    window.iconbitmap(resource_path("ico\MS.ico"))
    window.geometry("400x250")
    window.configure(bg="#f0f0f0")

    tk.Label(window, text="Đổi mật khẩu", font=("Arial", 14, "bold"), bg="#f0f0f0").pack(pady=10)

    frame = tk.Frame(window, bg="#f0f0f0")
    frame.pack(pady=10)

    tk.Label(frame, text="Mật khẩu cũ:", bg="#f0f0f0").grid(row=0, column=0, sticky="e", padx=5, pady=5)
    old_pass_entry = tk.Entry(frame, show="*", width=25)
    old_pass_entry.grid(row=0, column=1, padx=5, pady=5)

    tk.Label(frame, text="Mật khẩu mới:", bg="#f0f0f0").grid(row=1, column=0, sticky="e", padx=5, pady=5)
    new_pass_entry = tk.Entry(frame, show="*", width=25)
    new_pass_entry.grid(row=1, column=1, padx=5, pady=5)

    tk.Label(frame, text="Xác nhận lại:", bg="#f0f0f0").grid(row=2, column=0, sticky="e", padx=5, pady=5)
    confirm_pass_entry = tk.Entry(frame, show="*", width=25)
    confirm_pass_entry.grid(row=2, column=1, padx=5, pady=5)

    def update_password():
        old = old_pass_entry.get().strip()
        new = new_pass_entry.get().strip()
        confirm = confirm_pass_entry.get().strip()

        if not old or not new or not confirm:
            messagebox.showwarning("Thiếu thông tin", "Vui lòng điền đầy đủ các trường.")
            return

        if new != confirm:
            messagebox.showerror("Không khớp", "Mật khẩu mới và xác nhận không khớp.")
            return

        users = load_users()
        found = False

        for user in users["users"]:
            if user["username"] == current_username:
                if user["password"] != hash_password(old):
                    messagebox.showerror("Sai mật khẩu", "Mật khẩu cũ không đúng.")
                    return
                user["password"] = hash_password(new)
                found = True
                break

        if not found:
            messagebox.showerror("Lỗi", "Không tìm thấy người dùng.")
            return

        with open("User.json", "w", encoding="utf-8") as f:
            json.dump(users, f, indent=4, ensure_ascii=False)

        messagebox.showinfo("Thành công", "Đổi mật khẩu thành công.")
        window.destroy()

    tk.Button(window, text="Lưu thay đổi", bg="#4CAF50", fg="white", command=update_password).pack(pady=10)
    tk.Button(window, text="Đóng", command=window.destroy).pack()
#chạy tất cả các file
def load_all_user_files(file_tree):
    global current_username
    cr_file_name = f"{current_username}_orders.json"
    # Hiển thị cảnh báo xác nhận
    confirm = messagebox.askyesno(
        "Cảnh báo", 
        "Sau khi tải, các đơn hàng có thể bị thay đổi hoặc mất.\nBạn có chắc chắn muốn tiếp tục?"
    )
    if not confirm:
        return

    try:
        users = load_users()  # sửa lại từ load_user_files
        current_user = next((u for u in users["users"] if u["username"] == current_username), None)

        if not current_user or not current_user.get("data"):
            messagebox.showinfo("Thông báo", "Người dùng không có file dữ liệu nào")
            return

        all_orders = []
        order_counter = {}

        for filepath in current_user["data"]:
            if os.path.basename(filepath) == f"{current_username}_orders.json":
                continue
            try:
                manager = QuanLyDonHang(filepath)
                orders = manager.FileRead()

                for order in orders:
                    prefix = order.Ma[:5]
                    order_counter[prefix] = order_counter.get(prefix, 0) + 1
                    ma_moi = f"{prefix}{str(order_counter[prefix]).zfill(5)}"

                    new_order = DonHang(
                    ma_moi,
                    order.TenH,
                    order.NguoiGui,
                    order.NNhan,
                    order.cannang,
                    order.DVVC,
                    order.status,
                    order.From,
                    order.TO,
                    order.Pnumber1,
                    order.Pnumber2,
                    order.COD,
                    order.NgayDatHang
                )
                    all_orders.append(new_order)

            except Exception as e:
                print(f"Lỗi khi đọc file {filepath}: {e}")

    # Lưu file tổng hợp
        new_filename = f"{current_username}_orders.json"
        with open(new_filename, "w", encoding="utf-8") as f:
            json.dump([], f, indent=4, ensure_ascii=False)

        new_manager = QuanLyDonHang(new_filename)
        new_manager.save_orders(all_orders)

    # Cập nhật danh sách file người dùng
        if "data" not in current_user:
            current_user["data"] = []
        if new_filename not in current_user["data"]:
            current_user["data"].append(new_filename)

    # Ghi lại file User.json
        with open("User.json", "w", encoding="utf-8") as f:
            json.dump(users, f, indent=4, ensure_ascii=False)

        messagebox.showinfo(
        "Thành công", 
        f"Đã tạo file tổng hợp: {new_filename}\n"
        f"Tổng số đơn hàng: {len(all_orders)}"
    )


    except Exception as e:
        messagebox.showerror("Lỗi", f"Có lỗi xảy ra khi tạo file tổng hợp:\n{str(e)}")
def Kiem_tra_user_file(username):
    filename = f"{username}_orders.json"

    # 1. Tạo file nếu chưa có
    if not os.path.exists(filename):
        with open(filename, "w", encoding="utf-8") as f:
            json.dump([], f, indent=4, ensure_ascii=False)

    # 2. Đảm bảo trong User.json
    try:
        with open("User.json", "r", encoding="utf-8") as f:
            users = json.load(f)
    except FileNotFoundError:
        users = {"users": []}

    for user in users.get("users", []):
        if user.get("username") == username:
            if "data" not in user:
                user["data"] = []
            if filename not in user["data"]:
                user["data"].append(filename)
            break

    with open("User.json", "w", encoding="utf-8") as f:
        json.dump(users, f, indent=4, ensure_ascii=False)

    return filename
def new_user_file(tree):
    def create_file():
        filename = entry.get().strip()
        if not filename:
            messagebox.showwarning("Lỗi", "Vui lòng nhập tên file.")
            return

        if not filename.endswith(".json"):
            filename += ".json"

        if os.path.exists(filename):
            messagebox.showwarning("Tồn tại", f"File '{filename}' đã tồn tại.")
            return

        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump({}, f, ensure_ascii=False, indent=4)
                messagebox.showinfo("Thành công", f"Đã tạo file '{filename}' thành công.")
                users = load_users()

            for user in users["users"]:
                if user["username"] == current_username:
                    if "data" not in user:
                        user["data"] = []
                    if filename not in user["data"]:
                        user["data"].append(filename)
                    break
            else:
                messagebox.showerror("Lỗi", "Không tìm thấy người dùng hiện tại trong User.json")
                return

            # Ghi lại User.json
            with open("User.json", "w", encoding="utf-8") as f:
                json.dump(users, f, indent=4, ensure_ascii=False)

        # Bước 3: Làm mới lại giao diện TreeView
            load_user_files(tree)

            messagebox.showinfo("Thành công", f"Đã tạo file mới: {os.path.basename(filepath)}")
            window.destroy()
        except Exception as e:
            messagebox.showerror("Lỗi", f"Không thể tạo file: {e}")

    window = tk.Toplevel()
    window.title("Tạo file JSON")
    window.geometry("300x150")

    label = tk.Label(window, text="Nhập tên file JSON:")
    label.pack(pady=10)

    entry = tk.Entry(window, width=30)
    entry.pack()

    create_button = tk.Button(window, text="Tạo file", command=create_file)
    create_button.pack(pady=10)

# Gọi hàm này từ main window như sau:
# create_json_file_window()

#Giao diện chính
def MainProgram():
    global tree, total_label, delivered_label, delivering_label, shipping_label, canceled_label, Kho_van
    global current_file, current_user_role, current_username
    root = tk.Tk()
    root.title("Quản Lý Đơn Hàng")
    root.geometry("1200x600")
    root.iconbitmap(resource_path("ico\OIP.ico")) 

    # Tạo frame chứa thống kê và menu
    frame_stats = tk.Frame(root, pady=5)
    frame_stats.pack(fill="x", padx=10, pady=5)

    # Thêm các label thống kê bên trái
    total_label = tk.Label(frame_stats, text="Tổng đơn hàng: 0", font=("Arial", 12, "bold"))
    total_label.pack(side="left", padx=10)

    delivered_label = tk.Label(frame_stats, text="Đã giao: 0", font=("Arial", 12, "bold"), fg="green")
    delivered_label.pack(side="left", padx=10)

    delivering_label = tk.Label(frame_stats, text="Đang giao: 0", font=("Arial", 12, "bold"), fg="blue")
    delivering_label.pack(side="left", padx=10)
    
    Kho_van = tk.Label(frame_stats, text="Trong kho: 0", font=("Arial", 12, "bold"), fg="#7F00FF")
    Kho_van.pack(side="left", padx=10)

    shipping_label = tk.Label(frame_stats, text="Đang vận chuyển: 0", font=("Arial", 12, "bold"), fg="orange")  
    shipping_label.pack(side="left", padx=10)

    canceled_label = tk.Label(frame_stats, text="Đã hủy: 0", font=("Arial", 12, "bold"), fg="red")
    canceled_label.pack(side="left", padx=10)

    # Tạo nút menu bên phải
    menu_button = tk.Button(frame_stats, text="☰ ", font=("Arial", 12), bg="#f0f0f0")
    menu_button.pack(side="right", padx=10)

    # Tạo slide bar (ban đầu ẩn)
    slide_bar = tk.Frame(root, bg="#f0f0f0", width=170, bd=1, relief="sunken")
    slide_bar.pack_propagate(False)  # Ngăn không cho các widget con thay đổi kích thước frame
    slide_bar_visible = False

    # Thêm các lựa chọn vào slide bar
    options = ["Xuất đơn hàng ra file excel", "Quản lý", "Đổi mật khẩu","Trợ giúp", "Đăng xuất","Thoát"]
    for option in options:
        btn = tk.Button(slide_bar, text=option, width=15, relief="flat", anchor="w",
                       command=lambda opt=option: on_menu_select(opt))
        btn.pack(fill="x", padx=5, pady=3)

    # Hàm xử lý khi chọn menu
    def on_menu_select(option):
        if option == "Đăng xuất":
            Chose = messagebox.askyesno("Thông báo", f"Bạn sẽ đăng xuất khỏi ứng dụng, Xác nhận đăng xuất ?")
            if not Chose:
                return
            else:
                Reset_giatri()
                root.destroy()
                main()

        if option == "Quản lý":
            if current_user_role == "admin":
                admin_user_management()
            else:
                user_file_management()

        elif option == "Thoát":
            Chosse = messagebox.askyesno("Thông báo", f"Bạn có chắc muốn thoát chương trình không ?")
            if not Chosse:
                return
            root.destroy()
        elif option == "Trợ giúp":
            help_window()
        elif option == "Xuất đơn hàng ra file excel":
            if(current_user_role!="admin"):
                messagebox.showerror("Thông báo", "Chỉ quản trị viên mới có quyền xuất thông tin đơn hàng!")
                return
            else: export_to_excel()
        elif option == "Đổi mật khẩu":
            change_password()

        hide_slide_bar()

    # Hàm hiển thị slide bar
    def show_slide_bar():
        nonlocal slide_bar_visible
        slide_bar.place(x=root.winfo_width()-170, y=frame_stats.winfo_height()+10, 
                       height=root.winfo_height()-frame_stats.winfo_height()-10)
        slide_bar.lift()
        slide_bar_visible = True
    # Hàm ẩn slide bar
    def hide_slide_bar():
        nonlocal slide_bar_visible
        slide_bar.place_forget()
        slide_bar_visible = False
    # Hàm toggle slide bar
    def toggle_slide_bar():
        if slide_bar_visible:
            hide_slide_bar()
        else:
            show_slide_bar()

    menu_button.config(command=toggle_slide_bar)
    # Cập nhật vị trí slide bar khi thay đổi kích thước cửa sổ
    def on_resize(event):
        if slide_bar_visible:
            show_slide_bar()
    root.bind("<Configure>", on_resize)
    # Đóng slide bar khi click ra ngoài
    def close_slide_bar(event):
        if slide_bar_visible:
            # Kiểm tra nếu click không nằm trong slide bar
            if not (root.winfo_width()-170 <= event.x_root - root.winfo_rootx() <= root.winfo_width() and
                    frame_stats.winfo_height()+10 <= event.y_root - root.winfo_rooty() <= root.winfo_height()):
                hide_slide_bar()
    root.bind("<Button-1>", close_slide_bar)
    frame_buttons = tk.Frame(root)
    frame_buttons.pack(fill="x", padx=10, pady=5)
    tk.Button(frame_buttons, text="    Thống kê     ", bg="#00994C",command=lambda: ThongKe()).pack(side="left", padx=5)
    tk.Button(frame_buttons, text="Xóa", command=delete_selected_order, bg="#FF6666").pack(side="right", padx=5)
    tk.Button(frame_buttons, text="Chỉnh sửa", command=edit_selected_order, bg="#E28C6D").pack(side="right", padx=5)
    
    tk.Button(frame_buttons, text="   Thêm mới    ", bg="#ADCDDE", command=lambda: open_add_order_window(root, tree, order_manager)).pack(side="left", padx=5)
    tk.Button(frame_buttons, text="Lọc", bg="#FFCCCC",command=lambda: Loc_San_Pham(root, tree, order_manager)).pack(side="right", padx=5)
    tk.Button(frame_buttons, text="Xem chi tiết", command=show_selected_order, bg="#66FFB2").pack(side="right", padx=5)
    tk.Button(frame_buttons, text="Cập nhật lại", command=load_orders, bg="#99CCFF").pack(side="right", padx=5)
    sr = tk.Entry(frame_buttons, width=40)
    tk.Button(frame_buttons, text="Tìm kiếm", bg="#FFFF33", command= lambda:search_orders(sr)).pack(side="left", padx=5)
    sr.pack(side="left",padx=5)
    sr.insert(0,"Nhập tên, mã đơn hàng để tìm kiếm")
    sr.bind("<FocusIn>", lambda event: On_entry_click(event, sr, "Nhập tên, mã đơn hàng để tìm kiếm"))
    sr.bind("<FocusOut>", lambda event: On_focus_out(event, sr, "Nhập tên, mã đơn hàng để tìm kiếm"))
   
    frame_table = tk.Frame(root)
    frame_table.pack(fill="both", expand=True, padx=10, pady=5)

    scrollbar = ttk.Scrollbar(frame_table, orient="vertical")

    tree = ttk.Treeview(
        frame_table,
        columns=("STT", "Mã đơn hàng", "Tên đơn hàng", "Người nhận", "Trạng thái", "Địa chỉ nhận hàng", "Ngày đặt hàng"),
        show="headings",
        yscrollcommand=scrollbar.set
    )
    for col in tree["columns"]:
        tree.heading(col, text=col)

    tree.heading("Mã đơn hàng", text="Mã đơn hàng", command=sort_by_order_code)
    tree.heading("Ngày đặt hàng", text="Ngày đặt hàng ", command=sort_by_order_date)
    tree.heading("Trạng thái", text="Trạng thái", command=sort_by_status)

    scrollbar.config(command=tree.yview)
    scrollbar.pack(side="right", fill="y")

    tree.heading("STT", text="STT")
    tree.column("STT", width=50, anchor="center")

    for col, width in [("Mã đơn hàng", 120), ("Tên đơn hàng", 150), ("Người nhận", 120), 
                      ("Trạng thái", 100), ("Địa chỉ nhận hàng", 150), ("Ngày đặt hàng", 150)]:
        tree.heading(col, text=col)
        tree.column(col, width=width)

    tree.pack(fill="both", expand=True)
    if current_user_role == "admin":
        current_file = "all_oders.json"
        messagebox.showinfo("Thông báo", "Đã tải tất cả đơn hàng của người dùng")
        load_orders()
    load_orders()
    root.mainloop()
def open_order_management():
    global roo
    roo.destroy()
    MainProgram()
def on_entry_click(event, entry, default_text, is_password=False):
    if entry.get() == default_text:
        entry.delete(0, "end")
        entry.config(fg="black", show="*" if is_password else "")
def on_focus_out(event, entry, default_text, is_password=False):
    if entry.get() == "":
        entry.insert(0, default_text)
        entry.config(fg="grey", show="" if not is_password else "*")
#Hàm đọc file user
def load_users():
    try:
        with open("User.json", "r", encoding="utf-8") as f:
            data = json.load(f)
            if not isinstance(data, dict) or "users" not in data:
                # Tạo cấu trúc mới nếu file không đúng định dạng
                data = {"users": []}
                # Thêm admin mặc định
                data["users"].append({
                    "username": "admin",
                    "password": hash_password("1234"),  # Mã hóa mật khẩu
                    "email":"thangcuoi1984a@gmail.com",
                    "role": "admin",
                    "status": "active",
                    "data": ["all_oders.json"]
                })
                with open("User.json", "w", encoding="utf-8") as f:
                    json.dump(data, f, indent=4, ensure_ascii=False)
            return data
    except (FileNotFoundError, json.JSONDecodeError):
        # Tạo file mới nếu không tồn tại hoặc lỗi
        data = {
            "users": [{
                "username": "admin",
                "password": hash_password("1234"),
                "role": "admin",
                "email":"thangcuoi1984a@gmail.com",
                "status": "active",
                "data": ["all_oders.json"]
            }]
        }
        with open("User.json", "w", encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
        return data
def hash_password(password):
    """Mã hóa mật khẩu bằng SHA-256"""
    return hashlib.sha256(password.encode()).hexdigest()
#Xác minh mật khẩu
def verify_password(input_password, hashed_password):
    """Xác thực mật khẩu"""
    return hash_password(input_password) == hashed_password
def open_register_window():
   def register_user():
    username = entry_username.get().strip()
    password = entry_password.get().strip()
    confirm = entry_confirm.get().strip()
    email = entry_email.get().strip()

    if not username or not password or not confirm:
        messagebox.showwarning("Lỗi", "Vui lòng nhập đầy đủ thông tin.")
        return

    if password != confirm:
        messagebox.showerror("Lỗi", "Mật khẩu nhập lại không khớp!")
        return
    if not is_valid_email(email):
        messagebox.showerror("Lỗi", "Email không hợp lệ!")
        return

    users_data = load_users()  # Đây là dictionary chứa key "users"
    users = users_data.get("users", [])  # Lấy list users từ dictionary

    # Kiểm tra username đã tồn tại chưa
    if any(user["username"] == username for user in users):
        messagebox.showerror("Lỗi", "Tên người dùng đã tồn tại.")
        return

    # Tạo user mới
    new_user = {
        "username": username,
        "password": hash_password(password),
        "role": "user",
        "created_at": datetime.datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "status": "active",
        "data": []  # Khởi tạo với danh sách file trống
    }

    # Thêm user mới vào danh sách
    users.append(new_user)
    users_data["users"] = users  # Cập nhật lại dictionary

    # Lưu lại file
    with open("User.json", "w", encoding="utf-8") as f:
        json.dump(users_data, f, indent=4, ensure_ascii=False)

    messagebox.showinfo("Thành công", "Đăng ký thành công. Bạn có thể đăng nhập.")
    reg_win.destroy()
   reg_win = tk.Toplevel()
   reg_win.title("Đăng ký tài khoản")
   reg_win.geometry("450x250")
  

   tk.Label(reg_win, text="Vui lòng điền thông tin để đăng ký", font=("Arial", 10, "bold")).pack(pady=10)

   frame_inputs = tk.Frame(reg_win)
   frame_inputs.pack(pady=5)

    # Username
   tk.Label(frame_inputs, text="Tên đăng nhập:", font=("Arial", 10)).grid(row=0, column=0, padx=10, pady=5, sticky="e")
   entry_username = tk.Entry(frame_inputs, font=("Arial", 12), width=30)
   entry_username.grid(row=0, column=1, pady=5)

    # Password
   tk.Label(frame_inputs, text="Mật khẩu:", font=("Arial", 10)).grid(row=1, column=0, padx=10, pady=5, sticky="e")
   entry_password = tk.Entry(frame_inputs, font=("Arial", 12), width=30, show="*")
   entry_password.grid(row=1, column=1, pady=5)

    # Confirm password
   tk.Label(frame_inputs, text="Nhập lại mật khẩu:", font=("Arial", 10)).grid(row=2, column=0, padx=10, pady=5, sticky="e")
   entry_confirm = tk.Entry(frame_inputs, font=("Arial", 12), width=30, show="*")
   entry_confirm.grid(row=2, column=1, pady=5)
   # email
   tk.Label(frame_inputs, text="Email liên lạc:", font=("Arial", 10)).grid(row=3, column=0, padx=10, pady=5, sticky="e")
   entry_email = tk.Entry(frame_inputs, font=("Arial", 12), width=30)
   entry_email.grid(row=3, column=1, pady=5)
   

   tk.Button(reg_win, text="Đăng ký", command=register_user, bg="#4CAF50", fg="white",width=10).pack(pady=15)
def login(event=None):
    global current_user_role, current_username
    current_file = None
    username = entry_username.get()
    password = entry_password.get()
    users = load_users()

    for user in users.get("users", []):
        
        if user["username"] == username and verify_password(password, user["password"]):
            current_user_role = user["role"]
            current_username = username
            messagebox.showinfo("Thông báo", f"Đăng nhập thành công!\nChào mừng {username}")
            
            if current_user_role == "admin":
                create_admin_file()
                current_file = "all_orders.json"
            elif current_user_role == "user":
                print ("user")
                current_file = Kiem_tra_user_file(current_username)
                print("current_file:", current_file)
                order_manager = QuanLyDonHang(current_file)
            if user.get("status") != "active":
                messagebox.showerror("Tài khoản bị khóa", "Tài khoản của bạn đang bị khóa hoặc vô hiệu hóa, liên hệ quản trị viên để biết thêm thông tin hoặc mở khóa .")
                return
            open_order_management()
            return
    messagebox.showerror("Lỗi", "Tên người dùng hoặc mật khẩu không đúng!")
#tạo file admin  
def create_admin_file():
 
    """Tạo file mặc định của admin"""
    all_orders = []
    users = load_users()
    for user in users["users"]:
        if user["role"] == "admin":
            continue
        for filepath in user.get("data", []):
            try:
                manager = QuanLyDonHang(filepath)
                orders = manager.FileRead()
                if not orders: continue
                for order in orders:
                    od = order.to_dict()
                    od["_username"] = user["username"]
                    od["_source_file"] = filepath
                    all_orders.append(od)
            except Exception as e:
                print(f"Lỗi đọc file {filepath}: {e}")

    # Tạo file nếu chưa có
    with open("all_oders.json", "w", encoding="utf-8") as f:
        json.dump(all_orders, f, indent=4, ensure_ascii=False)
def update_order_in_user_file(edited_order_dict):
    """Cập nhật đơn hàng đã chỉnh sửa từ admin.json về file user gốc"""
    username = edited_order_dict.get("_username")
    source_file = edited_order_dict.get("_source_file")
    if not username or not source_file:
        messagebox.showerror("Lỗi", "Không thể xác định người dùng hoặc file nguồn.")
        return

    try:
        manager = QuanLyDonHang(source_file)
        orders = manager.FileRead()
        for i, order in enumerate(orders):
            if order.Ma == edited_order_dict["Mã đơn hàng"]:
                # Cập nhật lại đơn hàng
                orders[i] = DonHang(
                    edited_order_dict["Mã đơn hàng"],
                    edited_order_dict["Tên đơn hàng"],
                    edited_order_dict["Tên shop"],
                    edited_order_dict["Tên người nhận"],
                    edited_order_dict["Cân nặng (kg)"],
                    edited_order_dict["Đơn vị vận chuyển"],
                    edited_order_dict["Trạng thái"],
                    edited_order_dict["Địa chỉ lấy hàng"],
                    edited_order_dict["Địa chỉ nhận hàng"],
                    edited_order_dict["SĐT người nhận"],
                    edited_order_dict["SĐT người gửi"],
                    edited_order_dict["COD"],
                    edited_order_dict["Ngày đặt hàng"]
                )
                manager.save_orders(orders)
                messagebox.showinfo("Thành công", "Đã cập nhật đơn hàng vào file người dùng.")
                return
        messagebox.showwarning("Không tìm thấy", "Không tìm thấy đơn hàng để cập nhật.")
    except Exception as e:
        messagebox.showerror("Lỗi", f"Không thể cập nhật đơn hàng!\n{e}")
def QuenMatKhau():
    messagebox.showinfo("Thông báo", "Vui lòng liên hệ với quản trị viên để lấy lại mật khẩu")
def Login():
    global roo, entry_username, entry_password
    
    roo = tk.Tk()
    roo.title("Đăng nhập")
    roo.geometry("450x250+500+250")
    roo.resizable(False, False)
    roo.iconbitmap(resource_path("ico\OIP.ico"))

    label_title = tk.Label(roo, text="Vui lòng đăng nhập để vào giao diện quản lý", font=("Arial", 10, "bold"))
    label_title.pack(pady=10)

    frame_inputs = tk.Frame(roo)
    frame_inputs.pack(pady=5)

    label_username = tk.Label(frame_inputs, text="Tên đăng nhập:", font=("Arial", 10))
    label_username.grid(row=0, column=0, padx=10, pady=5, sticky="e")
    entry_username = tk.Entry(frame_inputs, font=("Arial", 12), width=30, fg="grey")
    entry_username.insert(0, "Nhập tên người dùng")
    entry_username.bind("<FocusIn>", lambda event: on_entry_click(event, entry_username, "Nhập tên người dùng"))
    entry_username.bind("<FocusOut>", lambda event: on_focus_out(event, entry_username, "Nhập tên người dùng"))
    entry_username.grid(row=0, column=1, pady=5)

    label_password = tk.Label(frame_inputs, text="Mật khẩu:", font=("Arial", 10))
    label_password.grid(row=1, column=0, padx=10, pady=5, sticky="e")
    entry_password = tk.Entry(frame_inputs, font=("Arial", 12), width=30, fg="grey")
    entry_password.insert(0, "Nhập password")
    entry_password.bind("<FocusIn>", lambda event: on_entry_click(event, entry_password, "Nhập password", True))
    entry_password.bind("<FocusOut>", lambda event: on_focus_out(event, entry_password, "Nhập password", True))
    entry_password.grid(row=1, column=1, pady=5)

    remember_var = tk.BooleanVar()
  
    frame_links = tk.Frame(roo)
    frame_links.pack(pady=5, fill="x", padx=50)

    # Đăng ký 
    btn_register = tk.Label(frame_links, text="Chưa có tài khoản?", font=("Arial", 10), fg="blue", cursor="hand2")
    btn_register.pack(side="left")
    btn_register.bind("<Button-1>", lambda e: open_register_window())

    # Quên mật khẩu 
    btn_forgot = tk.Label(frame_links, text="Quên mật khẩu?", font=("Arial", 10), fg="blue", cursor="hand2")
    btn_forgot.pack(side="right")
    btn_forgot.bind("<Button-1>", lambda e: QuenMatKhau())


    btn_login = tk.Button(roo, text="Đăng nhập", font=("Arial", 12), bg="#2C67F2", fg="white", width=20, height=2, command=login)
    
    btn_login.pack(pady=20)
    roo.bind("<Return>", login)
    roo.mainloop()
def main():
        Login()
if __name__ == "__main__": 
    main()



